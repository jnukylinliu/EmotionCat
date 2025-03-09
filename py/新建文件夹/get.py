import logging
from selenium import webdriver
from bs4 import BeautifulSoup
import re
import time
from openpyxl import Workbook

# 设置 logging，关闭 Selenium 的日志输出
logging.basicConfig(level=logging.ERROR)

# 设置 WebDriver（以 Chrome 为例）
options = webdriver.ChromeOptions()
options.add_argument('--headless')  # 无头模式，不显示浏览器窗口
driver = webdriver.Chrome(options=options)

# 目标网页 URL
url = "https://stock.10jqka.com.cn/fupan/"

# 加载网页
driver.get(url)

# 等待页面加载完成
time.sleep(3)  # 根据网络情况调整等待时间，确保JavaScript加载完成

# 获取网页源代码
html = driver.page_source

# 解析网页内容
soup = BeautifulSoup(html, 'html.parser')

# 查找包含“全市场成交额”文本的元素
text_elements = soup.find_all(string=re.compile(r'全市场成交额.*?\d+亿元'))

# 提取数字部分
成交额 = None
for text in text_elements:
    # 使用正则提取数字
    match = re.search(r'(\d+)亿元', text)
    if match:
        成交额 = match.group(1)
        print(成交额)  # 只输出数字部分
        break

# 关闭浏览器
driver.quit()

# 如果找到了成交额，保存到 Excel 文件
if 成交额:
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '成交额'  # 在A1单元格中写入表头
    ws['A2'] = 成交额   # 在A2单元格中写入成交额的值

    # 保存工作簿到 Excel 文件
    wb.save("成交额.xlsx")
    print("成交额已保存到 Excel 文件！")
else:
    print("未找到成交额！")
